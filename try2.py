import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet




def add_subtotals(sheet, start_row, start_col, end_col):
    # Adaugarea a trei randuri goale deasupra


    # Calcularea subtotalurilor pentru fiecare coloană
    for col in range(start_col, end_col + 1):
        # Formula pentru suma
        formula = f"=SUM({get_column_letter(col)}6:{get_column_letter(col)}{sheet.max_row})"
        
        # Adaugarea formulei, formatare si font
        cell = sheet.cell(row=start_row, column=col)
        cell.value = formula
        cell.number_format = '#,##0.00'
        cell.font = Font(bold=True)

# Funcție pentru a parsa XML-ul și a extrage informațiile relevante
def parse_xml_to_dict(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()
    
    namespaces = {'nsSAFT': 'mfp:anaf:dgti:d406:declaratie:v1'}
    
    accounts = []
    customers = []
    suppliers=[]
    invoices=[]
    invoices_sales=[]
    payments=[]
    je=[]
    
    # Extragerea datelor din secțiunea <Account>
    for account in root.findall('.//nsSAFT:Account', namespaces):
        account_data = {
            'AccountID': account.find('nsSAFT:AccountID', namespaces).text if account.find('nsSAFT:AccountID', namespaces) is not None else '',
            'AccountDescription': account.find('nsSAFT:AccountDescription', namespaces).text if account.find('nsSAFT:AccountDescription', namespaces) is not None else '',
            'AccountType': account.find('nsSAFT:AccountType', namespaces).text if account.find('nsSAFT:AccountType', namespaces) is not None else '',
            'OpeningDebitBalance': float(account.find('nsSAFT:OpeningDebitBalance', namespaces).text) if account.find('nsSAFT:OpeningDebitBalance', namespaces) is not None else 0,
            'OpeningCreditBalance': float(account.find('nsSAFT:OpeningCreditBalance', namespaces).text) if account.find('nsSAFT:OpeningCreditBalance', namespaces) is not None else 0,
            'ClosingDebitBalance': float(account.find('nsSAFT:ClosingDebitBalance', namespaces).text) if account.find('nsSAFT:ClosingDebitBalance', namespaces) is not None else 0,
            'ClosingCreditBalance': float(account.find('nsSAFT:ClosingCreditBalance', namespaces).text) if account.find('nsSAFT:ClosingCreditBalance', namespaces) is not None else 0
        }
        accounts.append(account_data)
    
    # Extragerea datelor din secțiunea <Customer>
  
    for customer in root.findall('.//nsSAFT:Customer', namespaces):
        company_structure = customer.find('nsSAFT:CompanyStructure', namespaces)
        customer_data = {
            'CustomerID': customer.find('nsSAFT:CustomerID', namespaces).text,
            'AccountID': customer.find('nsSAFT:AccountID', namespaces).text,
            'OpeningDebitBalance': float(customer.find('nsSAFT:OpeningDebitBalance', namespaces).text) if customer.find('nsSAFT:OpeningDebitBalance', namespaces) is not None else 0,
            'ClosingDebitBalance': float(customer.find('nsSAFT:ClosingDebitBalance', namespaces).text) if customer.find('nsSAFT:ClosingDebitBalance', namespaces) is not None else 0,
            'OpeningCreditBalance': float(customer.find('nsSAFT:OpeningCreditBalance', namespaces).text) if customer.find('nsSAFT:OpeningCreditBalance', namespaces) is not None else 0,
            'ClosingCreditBalance': float(customer.find('nsSAFT:ClosingCreditBalance', namespaces).text) if customer.find('nsSAFT:ClosingCreditBalance', namespaces) is not None else 0,
            'RegistrationNumber': company_structure.find('nsSAFT:RegistrationNumber', namespaces).text if company_structure.find('nsSAFT:RegistrationNumber', namespaces) is not None else '',
            'Name': company_structure.find('nsSAFT:Name', namespaces).text if company_structure.find('nsSAFT:Name', namespaces) is not None else '',
            'City': company_structure.find('.//nsSAFT:City', namespaces).text if company_structure.find('.//nsSAFT:City', namespaces) is not None else '',
            'Country': company_structure.find('.//nsSAFT:Country', namespaces).text if company_structure.find('.//nsSAFT:Country', namespaces) is not None else ''
        }
        customers.append(customer_data)

    for supplier in root.findall('.//nsSAFT:Supplier', namespaces):
        company_structure = supplier.find('nsSAFT:CompanyStructure', namespaces)
        supplier_data = {
            'SupplierID': supplier.find('nsSAFT:SupplierID', namespaces).text,
            'AccountID': supplier.find('nsSAFT:AccountID', namespaces).text,
            'OpeningDebitBalance': float(supplier.find('nsSAFT:OpeningDebitBalance', namespaces).text) if supplier.find('nsSAFT:OpeningDebitBalance', namespaces) is not None else 0,
            'ClosingDebitBalance': float(supplier.find('nsSAFT:ClosingDebitBalance', namespaces).text) if supplier.find('nsSAFT:ClosingDebitBalance', namespaces) is not None else 0,
            'OpeningCreditBalance': float(supplier.find('nsSAFT:OpeningCreditBalance', namespaces).text) if supplier.find('nsSAFT:OpeningCreditBalance', namespaces) is not None else 0,
            'ClosingCreditBalance': float(supplier.find('nsSAFT:ClosingCreditBalance', namespaces).text) if supplier.find('nsSAFT:ClosingCreditBalance', namespaces) is not None else 0,
            'RegistrationNumber': company_structure.find('nsSAFT:RegistrationNumber', namespaces).text if company_structure.find('nsSAFT:RegistrationNumber', namespaces) is not None else '',
            'Name': company_structure.find('nsSAFT:Name', namespaces).text if company_structure.find('nsSAFT:Name', namespaces) is not None else '',
            'City': company_structure.find('.//nsSAFT:City', namespaces).text if company_structure.find('.//nsSAFT:City', namespaces) is not None else '',
            'Country': company_structure.find('.//nsSAFT:Country', namespaces).text if company_structure.find('.//nsSAFT:Country', namespaces) is not None else ''
        }
        suppliers.append(supplier_data)
    for invoice in root.findall('.//nsSAFT:PurchaseInvoices/nsSAFT:Invoice', namespaces):
        invoice_data = {
            'InvoiceNo': invoice.find('nsSAFT:InvoiceNo', namespaces).text if invoice.find('nsSAFT:InvoiceNo', namespaces) is not None else '',
            'SupplierID': invoice.find('.//nsSAFT:SupplierInfo/nsSAFT:SupplierID', namespaces).text if invoice.find('.//nsSAFT:SupplierInfo/nsSAFT:SupplierID', namespaces) is not None else '',
            'BillingCity': invoice.find('.//nsSAFT:BillingAddress/nsSAFT:City', namespaces).text if invoice.find('.//nsSAFT:BillingAddress/nsSAFT:City', namespaces) is not None else '',
            'BillingCountry': invoice.find('.//nsSAFT:BillingAddress/nsSAFT:Country', namespaces).text if invoice.find('.//nsSAFT:BillingAddress/nsSAFT:Country', namespaces) is not None else '',
            'AccountID': invoice.find('nsSAFT:AccountID', namespaces).text if invoice.find('nsSAFT:AccountID', namespaces) is not None else '',
            'InvoiceDate': invoice.find('nsSAFT:InvoiceDate', namespaces).text if invoice.find('nsSAFT:InvoiceDate', namespaces) is not None else '',
            'InvoiceType': invoice.find('nsSAFT:InvoiceType', namespaces).text if invoice.find('nsSAFT:InvoiceType', namespaces) is not None else '',
            'SelfBillingIndicator': invoice.find('nsSAFT:SelfBillingIndicator', namespaces).text if invoice.find('nsSAFT:SelfBillingIndicator', namespaces) is not None else '',
            'InvoiceLines': []
        }
        
        lines_data = []
        for line in invoice.findall('.//nsSAFT:InvoiceLine', namespaces):
            line_data = {
                'InvoiceNo': invoice_data['InvoiceNo'],
                'AccountID': line.find('nsSAFT:AccountID', namespaces).text if line.find('nsSAFT:AccountID', namespaces) is not None else '',
                'Quantity': float(line.find('nsSAFT:Quantity', namespaces).text) if line.find('nsSAFT:Quantity', namespaces) is not None else 0,
                'UnitPrice': float(line.find('nsSAFT:UnitPrice', namespaces).text) if line.find('nsSAFT:UnitPrice', namespaces) is not None else 0,
                'TaxPointDate': line.find('nsSAFT:TaxPointDate', namespaces).text if line.find('nsSAFT:TaxPointDate', namespaces) is not None else '',
                'Description': line.find('nsSAFT:Description', namespaces).text if line.find('nsSAFT:Description', namespaces) is not None else '',
                'InvoiceLineAmount': float(line.find('.//nsSAFT:InvoiceLineAmount/nsSAFT:Amount', namespaces).text) if line.find('.//nsSAFT:InvoiceLineAmount/nsSAFT:Amount', namespaces) is not None else 0,
                'CurrencyCode': line.find('.//nsSAFT:InvoiceLineAmount/nsSAFT:CurrencyCode', namespaces).text if line.find('.//nsSAFT:InvoiceLineAmount/nsSAFT:CurrencyCode', namespaces) is not None else '',
                'CurrencyAmount': float(line.find('.//nsSAFT:InvoiceLineAmount/nsSAFT:CurrencyAmount', namespaces).text) if line.find('.//nsSAFT:InvoiceLineAmount/nsSAFT:CurrencyAmount', namespaces) is not None else 0,
                'DebitCreditIndicator': line.find('nsSAFT:DebitCreditIndicator', namespaces).text if line.find('nsSAFT:DebitCreditIndicator', namespaces) is not None else '',
                'TaxType': line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxType', namespaces).text if line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxType', namespaces) is not None else '',
                'TaxCode': line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxCode', namespaces).text if line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxCode', namespaces) is not None else '',
                'TaxAmount': float(line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxAmount/nsSAFT:Amount', namespaces).text) if line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxAmount/nsSAFT:Amount', namespaces) is not None else 0,
                'TaxCurrencyCode': line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxAmount/nsSAFT:CurrencyCode', namespaces).text if line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxAmount/nsSAFT:CurrencyCode', namespaces) is not None else '',
                'TaxCurrencyAmount': float(line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxAmount/nsSAFT:CurrencyAmount', namespaces).text) if line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxAmount/nsSAFT:CurrencyAmount', namespaces) is not None else 0
            }
            lines_data.append(line_data)
        
        invoice_df = pd.DataFrame(lines_data)
        invoice_df['BillingCity'] = invoice_data['BillingCity']
        invoice_df['SupplierID'] = invoice_data['SupplierID']
        invoice_df['BillingCountry'] = invoice_data['BillingCountry']
        invoices.append(invoice_df)
    invoices=pd.concat(invoices)
        
   

    for invoice_sales in root.findall('.//nsSAFT:SalesInvoices/nsSAFT:Invoice', namespaces):
        invoice_data_sales = {
            'InvoiceNo': invoice_sales.find('nsSAFT:InvoiceNo', namespaces).text if invoice_sales.find('nsSAFT:InvoiceNo', namespaces) is not None else '',
            'CustomerID': invoice_sales.find('.//nsSAFT:CustomerInfo/nsSAFT:CustomerID', namespaces).text if invoice_sales.find('.//nsSAFT:CustomerInfo/nsSAFT:CustomerID', namespaces) is not None else '',
            'BillingCity': invoice_sales.find('.//nsSAFT:BillingAddress/nsSAFT:City', namespaces).text if invoice_sales.find('.//nsSAFT:BillingAddress/nsSAFT:City', namespaces) is not None else '',
            'BillingCountry': invoice_sales.find('.//nsSAFT:BillingAddress/nsSAFT:Country', namespaces).text if invoice_sales.find('.//nsSAFT:BillingAddress/nsSAFT:Country', namespaces) is not None else '',
            'AccountID': invoice_sales.find('nsSAFT:AccountID', namespaces).text if invoice_sales.find('nsSAFT:AccountID', namespaces) is not None else '',
            'InvoiceDate': invoice_sales.find('nsSAFT:InvoiceDate', namespaces).text if invoice_sales.find('nsSAFT:InvoiceDate', namespaces) is not None else '',
            'InvoiceType': invoice_sales.find('nsSAFT:InvoiceType', namespaces).text if invoice_sales.find('nsSAFT:InvoiceType', namespaces) is not None else '',
            'SelfBillingIndicator': invoice_sales.find('nsSAFT:SelfBillingIndicator', namespaces).text if invoice_sales.find('nsSAFT:SelfBillingIndicator', namespaces) is not None else '',
            'InvoiceLines': [] 
            }
        # print(invoice_data_sales)
        lines_data = []
        for line in invoice_sales.findall('.//nsSAFT:InvoiceLine', namespaces):
            line_data = {
                'InvoiceNo': invoice_data_sales['InvoiceNo'],
                'AccountID': line.find('nsSAFT:AccountID', namespaces).text if line.find('nsSAFT:AccountID', namespaces) is not None else '',
                'Quantity': float(line.find('nsSAFT:Quantity', namespaces).text) if line.find('nsSAFT:Quantity', namespaces) is not None else 0,
                'UnitPrice': float(line.find('nsSAFT:UnitPrice', namespaces).text) if line.find('nsSAFT:UnitPrice', namespaces) is not None else 0,
                'TaxPointDate': line.find('nsSAFT:TaxPointDate', namespaces).text if line.find('nsSAFT:TaxPointDate', namespaces) is not None else '',
                'Description': line.find('nsSAFT:Description', namespaces).text if line.find('nsSAFT:Description', namespaces) is not None else '',
                'InvoiceLineAmount': float(line.find('.//nsSAFT:InvoiceLineAmount/nsSAFT:Amount', namespaces).text) if line.find('.//nsSAFT:InvoiceLineAmount/nsSAFT:Amount', namespaces) is not None else 0,
                'CurrencyCode': line.find('.//nsSAFT:InvoiceLineAmount/nsSAFT:CurrencyCode', namespaces).text if line.find('.//nsSAFT:InvoiceLineAmount/nsSAFT:CurrencyCode', namespaces) is not None else '',
                'CurrencyAmount': float(line.find('.//nsSAFT:InvoiceLineAmount/nsSAFT:CurrencyAmount', namespaces).text) if line.find('.//nsSAFT:InvoiceLineAmount/nsSAFT:CurrencyAmount', namespaces) is not None else 0,
                'DebitCreditIndicator': line.find('nsSAFT:DebitCreditIndicator', namespaces).text if line.find('nsSAFT:DebitCreditIndicator', namespaces) is not None else '',
                'TaxType': line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxType', namespaces).text if line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxType', namespaces) is not None else '',
                'TaxCode': line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxCode', namespaces).text if line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxCode', namespaces) is not None else 0,
                'TaxAmount': float(line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxAmount/nsSAFT:Amount', namespaces).text) if line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxAmount/nsSAFT:Amount', namespaces) is not None else '',
                'TaxCurrencyCode': line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxAmount/nsSAFT:CurrencyCode', namespaces).text if line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxAmount/nsSAFT:CurrencyCode', namespaces) is not None else '',
                'TaxCurrencyAmount': float(line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxAmount/nsSAFT:CurrencyAmount', namespaces).text) if line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxAmount/nsSAFT:CurrencyAmount', namespaces) is not None else 0
            }
            lines_data.append(line_data)
        invoice_df = pd.DataFrame(lines_data)
        invoice_df['BillingCity'] = invoice_data_sales['BillingCity']
        invoice_df['CustomerID'] = invoice_data_sales['CustomerID']
        invoice_df['BillingCountry'] = invoice_data_sales['BillingCountry']
        invoices_sales.append(invoice_df)
        
        
    invoices_sales=pd.concat(invoices_sales)


    for payment in root.findall('.//nsSAFT:Payment', namespaces):
        payment_data = {
        'PaymentRefNo': payment.find('nsSAFT:PaymentRefNo', namespaces).text if payment.find('nsSAFT:PaymentRefNo', namespaces) is not None else '',
        'TransactionDate': payment.find('nsSAFT:TransactionDate', namespaces).text if payment.find('nsSAFT:TransactionDate', namespaces) is not None else '',
        'PaymentMethod': payment.find('nsSAFT:PaymentMethod', namespaces).text if payment.find('nsSAFT:PaymentMethod', namespaces) is not None else '',
        'Description': payment.find('nsSAFT:Description', namespaces).text if payment.find('nsSAFT:Description', namespaces) is not None else '',
        'PaymentLines': []
    }

        lines_data = []
        for line in payment.findall('.//nsSAFT:PaymentLine', namespaces):
            line_data = {
                'PaymentRefNo': payment_data['PaymentRefNo'],
                'AccountID': line.find('nsSAFT:AccountID', namespaces).text if line.find('nsSAFT:AccountID', namespaces) is not None else '',
                'CustomerID': line.find('nsSAFT:CustomerID', namespaces).text if line.find('nsSAFT:CustomerID', namespaces) is not None else '',
                'SupplierID': line.find('nsSAFT:SupplierID', namespaces).text if line.find('nsSAFT:SupplierID', namespaces) is not None else '',
                'DebitCreditIndicator': line.find('nsSAFT:DebitCreditIndicator', namespaces).text if line.find('nsSAFT:DebitCreditIndicator', namespaces) is not None else '',
                'PaymentLineAmount': float(line.find('.//nsSAFT:PaymentLineAmount/nsSAFT:Amount', namespaces).text) if line.find('.//nsSAFT:PaymentLineAmount/nsSAFT:Amount', namespaces) is not None else 0,
                'CurrencyCode': line.find('.//nsSAFT:PaymentLineAmount/nsSAFT:CurrencyCode', namespaces).text if line.find('.//nsSAFT:PaymentLineAmount/nsSAFT:CurrencyCode', namespaces) is not None else '',
                'CurrencyAmount': float(line.find('.//nsSAFT:PaymentLineAmount/nsSAFT:CurrencyAmount', namespaces).text) if line.find('.//nsSAFT:PaymentLineAmount/nsSAFT:CurrencyAmount', namespaces) is not None else 0,
                'TaxType': line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxType', namespaces).text if line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxType', namespaces) is not None else '',
                'TaxCode': line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxCode', namespaces).text if line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxCode', namespaces) is not None else '',
                'TaxAmount': float(line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxAmount/nsSAFT:Amount', namespaces).text) if line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxAmount/nsSAFT:Amount', namespaces) is not None else 0,
                'TaxCurrencyCode': line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxAmount/nsSAFT:CurrencyCode', namespaces).text if line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxAmount/nsSAFT:CurrencyCode', namespaces) is not None else '',
                'TaxCurrencyAmount': float(line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxAmount/nsSAFT:CurrencyAmount', namespaces).text) if line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxAmount/nsSAFT:CurrencyAmount', namespaces) is not None else 0
            }
            lines_data.append(line_data)
    
        payments_df = pd.DataFrame(lines_data)
        payments_df['PaymentRefNo'] = payment_data['PaymentRefNo']
        payments_df['TransactionDate'] = payment_data['TransactionDate']
        payments_df['PaymentMethod'] = payment_data['PaymentMethod']
        payments_df['Description'] = payment_data['Description']
        payments.append(payments_df)
    payments=pd.concat(payments)
    for transaction in root.findall('.//nsSAFT:Transaction', namespaces):
        transaction_data = {
            'TransactionID': transaction.find('nsSAFT:TransactionID', namespaces).text if transaction.find('nsSAFT:TransactionID', namespaces) is not None else '',
            'TransactionDate': transaction.find('nsSAFT:TransactionDate', namespaces).text if transaction.find('nsSAFT:TransactionDate', namespaces) is not None else '',
            'Description': transaction.find('nsSAFT:Description', namespaces).text if transaction.find('nsSAFT:Description', namespaces) is not None else '',
            'CustomerID': transaction.find('nsSAFT:CustomerID', namespaces).text if transaction.find('nsSAFT:CustomerID', namespaces) is not None else '',
            'SupplierID': transaction.find('nsSAFT:SupplierID', namespaces).text if transaction.find('nsSAFT:SupplierID', namespaces) is not None else '',
            'TransactionLines': []
        }

        lines_data = []
        for line in transaction.findall('.//nsSAFT:TransactionLine', namespaces):
            if line.find('nsSAFT:DebitAmount', namespaces) is not None:
                amount_tag = 'nsSAFT:DebitAmount'
            else:
                amount_tag = 'nsSAFT:CreditAmount'
                
            line_data = {
                'TransactionID': transaction_data['TransactionID'],
                'AccountID': line.find('nsSAFT:AccountID', namespaces).text if line.find('nsSAFT:AccountID', namespaces) is not None else '',
                'CustomerID': line.find('nsSAFT:CustomerID', namespaces).text if line.find('nsSAFT:CustomerID', namespaces) is not None else '',
                'SupplierID': line.find('nsSAFT:SupplierID', namespaces).text if line.find('nsSAFT:SupplierID', namespaces) is not None else '',
                'DebitCreditIndicator': 'Debit' if amount_tag == 'nsSAFT:DebitAmount' else 'Credit',
                'Amount': float(line.find(f'.//{amount_tag}/nsSAFT:Amount', namespaces).text) if line.find(f'.//{amount_tag}/nsSAFT:Amount', namespaces) is not None else 0,
                'CurrencyCode': line.find(f'.//{amount_tag}/nsSAFT:CurrencyCode', namespaces).text if line.find(f'.//{amount_tag}/nsSAFT:CurrencyCode', namespaces) is not None else '',
                'CurrencyAmount': float(line.find(f'.//{amount_tag}/nsSAFT:CurrencyAmount', namespaces).text) if line.find(f'.//{amount_tag}/nsSAFT:CurrencyAmount', namespaces) is not None else 0,
                'TaxType': line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxType', namespaces).text if line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxType', namespaces) is not None else '',
                'TaxCode': line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxCode', namespaces).text if line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxCode', namespaces) is not None else '',
                'TaxAmount': float(line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxAmount/nsSAFT:Amount', namespaces).text) if line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxAmount/nsSAFT:Amount', namespaces) is not None else 0,
                'TaxCurrencyCode': line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxAmount/nsSAFT:CurrencyCode', namespaces).text if line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxAmount/nsSAFT:CurrencyCode', namespaces) is not None else '',
                'TaxCurrencyAmount': float(line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxAmount/nsSAFT:CurrencyAmount', namespaces).text) if line.find('.//nsSAFT:TaxInformation/nsSAFT:TaxAmount/nsSAFT:CurrencyAmount', namespaces) is not None else 0
            }
            lines_data.append(line_data)
        
        transaction_df = pd.DataFrame(lines_data)
        transaction_df['TransactionID'] = transaction_data['TransactionID']
        transaction_df['TransactionDate'] = transaction_data['TransactionDate']
        transaction_df['Description'] = transaction_data['Description']
        transaction_df['CustomerID'] = transaction_data['CustomerID']
        transaction_df['SupplierID'] = transaction_data['SupplierID']
        je.append(transaction_df)

    je = pd.concat(je, ignore_index=True)


    
    return accounts, customers, suppliers, invoices, invoices_sales, payments, je


def add_empty_rows_before_data(sheet, num_rows):
    for _ in range(num_rows):
        sheet.append(['' for _ in range(sheet.max_column)])



def save_to_excel(accounts, customers, suppliers, invoices, invoices_sales, payments, je, client_name, output_file):
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Salvarea datelor în fișierul Excel
        df_accounts = pd.DataFrame(accounts)
        df_accounts.to_excel(writer, sheet_name='TB', index=False)

        df_customers = pd.DataFrame(customers)
        df_customers.to_excel(writer, sheet_name='Customers', index=False)

        df_suppliers = pd.DataFrame(suppliers)
        df_suppliers.to_excel(writer, sheet_name='Suppliers', index=False)

        df_purchase_invoices = pd.DataFrame(invoices)
        df_purchase_invoices.to_excel(writer, sheet_name='Purchase Invoices', index=False)

        df_sales_invoices = pd.DataFrame(invoices_sales)
        df_sales_invoices.to_excel(writer, sheet_name='Sales Invoices', index=False)

        df_payments = pd.DataFrame(payments)
        df_payments.to_excel(writer, sheet_name='Payments', index=False)

        df_je = pd.DataFrame(je)
        df_je.to_excel(writer, sheet_name='GL', index=False)

        # Accesarea fiecărei foi pentru aplicarea formatării și adăugarea rândurilor goale
        for sheet_name in writer.sheets:
            sheet = writer.sheets[sheet_name]

            # Ascunde liniile grilei
            sheet.sheet_view.showGridLines = False

            # Adaugă trei rânduri goale înainte de date
        

            # Setează lățimea coloanelor pentru a se potrivi conținutului
            for column_cells in sheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                sheet.column_dimensions[column_cells[0].column].width = length + 2

            # Adaugă subtotaluri înainte de date
           

            # Aplică formatarea antetului de tabel
            for cell in sheet['1:1']:
                cell.fill = PatternFill(start_color="674ea7", end_color="674ea7", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Aplică marginile
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.border = Border(left=Side(border_style="thin", color="000000"),
                                         right=Side(border_style="thin", color="000000"),
                                         top=Side(border_style="thin", color="000000"),
                                         bottom=Side(border_style="thin", color="000000"))

def save_to_excel_Assets(accounts,assets,asset_transactions, output_file):
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Salvarea datelor în fișierul Excel
        df_accounts = pd.DataFrame(accounts)
        df_accounts.to_excel(writer, sheet_name='TB', index=False)

        df_Assets = pd.DataFrame(assets)
        df_Assets.to_excel(writer, sheet_name='Assets', index=False)

        df_AssetTransactions = pd.DataFrame(asset_transactions)
        df_AssetTransactions.to_excel(writer, sheet_name='AssetTransactions', index=False)

        # df_purchase_invoices = pd.DataFrame(invoices)
        # df_purchase_invoices.to_excel(writer, sheet_name='Purchase Invoices', index=False)

        # df_sales_invoices = pd.DataFrame(invoices_sales)
        # df_sales_invoices.to_excel(writer, sheet_name='Sales Invoices', index=False)

        # df_payments = pd.DataFrame(payments)
        # df_payments.to_excel(writer, sheet_name='Payments', index=False)

        # df_je = pd.DataFrame(je)
        # df_je.to_excel(writer, sheet_name='GL', index=False)

        # Accesarea fiecărei foi pentru aplicarea formatării și adăugarea rândurilor goale
        for sheet_name in writer.sheets:
            sheet = writer.sheets[sheet_name]

            # Ascunde liniile grilei
            sheet.sheet_view.showGridLines = False

            # Adaugă trei rânduri goale înainte de date
        

            # Setează lățimea coloanelor pentru a se potrivi conținutului
            for column_cells in sheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                sheet.column_dimensions[column_cells[0].column].width = length + 2

            # Adaugă subtotaluri înainte de date
           

            # Aplică formatarea antetului de tabel
            for cell in sheet['1:1']:
                cell.fill = PatternFill(start_color="674ea7", end_color="674ea7", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Aplică marginile
            for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
                for cell in row:
                    cell.border = Border(left=Side(border_style="thin", color="000000"),
                                         right=Side(border_style="thin", color="000000"),
                                         top=Side(border_style="thin", color="000000"),
                                         bottom=Side(border_style="thin", color="000000"))

        # create_summary_sheet(writer, client_name)

# def create_summary_sheet(writer, client_name):
#     # Create a new summary sheet
#     ws = writer.book.create_sheet(title="Summary")

#     # Set sheet properties
#     ws.sheet_properties.tabColor = "674ea7  # Color the tab
#     ws.sheet_view.showGridLines = False  # Hide gridlines

#     # Add client name
#     ws['A1'] = f"Client: {client_name}"
#     ws['A1'].font = Font(size=14, bold=True)

#     # Add title
#     ws['A3'] = "Summary of Reports"
#     ws['A3'].font = Font(size=12, bold=True)

#     # Add total amounts
#     row = 5
#     summary_data = {
#         'TB': ['D', 'E', 'F', 'G'],
#         'Customers': ['C', 'D', 'E'],
#         'Suppliers': ['C', 'D', 'E'],
#         'Purchase Invoices': ['C', 'D', 'E'],
#         'Sales Invoices': ['C', 'D', 'E'],
#         'Payments': ['C', 'D', 'E'],
#         'GL': ['C', 'D', 'E']
#     }
def parse_xmlAssets_to_dict(xml_file):
    tree = ET.parse(xml_file)
    root = tree.getroot()
    
    namespaces = {'nsSAFT': 'mfp:anaf:dgti:d406:declaratie:v1'}
    
    accounts = []
    # customers = []
    # suppliers=[]
    # invoices=[]
    # invoices_sales=[]
    # payments=[]
    # je=[]
    
    # Extragerea datelor din secțiunea <Account>
    for account in root.findall('.//nsSAFT:Account', namespaces):
        account_data = {
            'AccountID': account.find('nsSAFT:AccountID', namespaces).text if account.find('nsSAFT:AccountID', namespaces) is not None else '',
            'AccountDescription': account.find('nsSAFT:AccountDescription', namespaces).text if account.find('nsSAFT:AccountDescription', namespaces) is not None else '',
            'AccountType': account.find('nsSAFT:AccountType', namespaces).text if account.find('nsSAFT:AccountType', namespaces) is not None else '',
            'OpeningDebitBalance': float(account.find('nsSAFT:OpeningDebitBalance', namespaces).text) if account.find('nsSAFT:OpeningDebitBalance', namespaces) is not None else 0,
            'OpeningCreditBalance': float(account.find('nsSAFT:OpeningCreditBalance', namespaces).text) if account.find('nsSAFT:OpeningCreditBalance', namespaces) is not None else 0,
            'ClosingDebitBalance': float(account.find('nsSAFT:ClosingDebitBalance', namespaces).text) if account.find('nsSAFT:ClosingDebitBalance', namespaces) is not None else 0,
            'ClosingCreditBalance': float(account.find('nsSAFT:ClosingCreditBalance', namespaces).text) if account.find('nsSAFT:ClosingCreditBalance', namespaces) is not None else 0
        }
        accounts.append(account_data)



    assets = []

    for asset in root.findall('.//nsSAFT:Asset', namespaces):
        asset_data = {
            'AssetID': asset.find('nsSAFT:AssetID', namespaces).text if asset.find('nsSAFT:AssetID', namespaces) is not None else '',
            'AccountID': asset.find('nsSAFT:AccountID', namespaces).text if asset.find('nsSAFT:AccountID', namespaces) is not None else '',
            'Description': asset.find('nsSAFT:Description', namespaces).text if asset.find('nsSAFT:Description', namespaces) is not None else '',
            'DateOfAcquisition': asset.find('nsSAFT:DateOfAcquisition', namespaces).text if asset.find('nsSAFT:DateOfAcquisition', namespaces) is not None else '',
            'StartUpDate': asset.find('nsSAFT:StartUpDate', namespaces).text if asset.find('nsSAFT:StartUpDate', namespaces) is not None else '',
        }
        
        valuation = asset.find('nsSAFT:Valuations/nsSAFT:Valuation', namespaces)
        if valuation is not None:
            valuation_data = {
                'AssetValuationType': valuation.find('nsSAFT:AssetValuationType', namespaces).text if valuation.find('nsSAFT:AssetValuationType', namespaces) is not None else '',
                'ValuationClass': valuation.find('nsSAFT:ValuationClass', namespaces).text if valuation.find('nsSAFT:ValuationClass', namespaces) is not None else '',
                'AcquisitionAndProductionCostsBegin': float(valuation.find('nsSAFT:AcquisitionAndProductionCostsBegin', namespaces).text) if valuation.find('nsSAFT:AcquisitionAndProductionCostsBegin', namespaces) is not None else 0,
                'AcquisitionAndProductionCostsEnd': float(valuation.find('nsSAFT:AcquisitionAndProductionCostsEnd', namespaces).text) if valuation.find('nsSAFT:AcquisitionAndProductionCostsEnd', namespaces) is not None else 0,
                'InvestmentSupport': float(valuation.find('nsSAFT:InvestmentSupport', namespaces).text) if valuation.find('nsSAFT:InvestmentSupport', namespaces) is not None else 0,
                'AssetLifeYear': float(valuation.find('nsSAFT:AssetLifeYear', namespaces).text) if valuation.find('nsSAFT:AssetLifeYear', namespaces) is not None else 0,
                'AssetAddition': float(valuation.find('nsSAFT:AssetAddition', namespaces).text) if valuation.find('nsSAFT:AssetAddition', namespaces) is not None else 0,
                'Transfers': float(valuation.find('nsSAFT:Transfers', namespaces).text) if valuation.find('nsSAFT:Transfers', namespaces) is not None else 0,
                'AssetDisposal': float(valuation.find('nsSAFT:AssetDisposal', namespaces).text) if valuation.find('nsSAFT:AssetDisposal', namespaces) is not None else 0,
                'BookValueBegin': float(valuation.find('nsSAFT:BookValueBegin', namespaces).text) if valuation.find('nsSAFT:BookValueBegin', namespaces) is not None else 0,
                'DepreciationMethod': valuation.find('nsSAFT:DepreciationMethod', namespaces).text if valuation.find('nsSAFT:DepreciationMethod', namespaces) is not None else '',
                'DepreciationPercentage': float(valuation.find('nsSAFT:DepreciationPercentage', namespaces).text) if valuation.find('nsSAFT:DepreciationPercentage', namespaces) is not None else 0,
                'DepreciationForPeriod': float(valuation.find('nsSAFT:DepreciationForPeriod', namespaces).text) if valuation.find('nsSAFT:DepreciationForPeriod', namespaces) is not None else 0,
                'AppreciationForPeriod': float(valuation.find('nsSAFT:AppreciationForPeriod', namespaces).text) if valuation.find('nsSAFT:AppreciationForPeriod', namespaces) is not None else 0,
                'AccumulatedDepreciation': float(valuation.find('nsSAFT:AccumulatedDepreciation', namespaces).text) if valuation.find('nsSAFT:AccumulatedDepreciation', namespaces) is not None else 0,
                'BookValueEnd': float(valuation.find('nsSAFT:BookValueEnd', namespaces).text) if valuation.find('nsSAFT:BookValueEnd', namespaces) is not None else 0
            }
            
            extraordinary_depreciation = valuation.find('nsSAFT:ExtraordinaryDepreciationsForPeriod/nsSAFT:ExtraordinaryDepreciationForPeriod', namespaces)
            if extraordinary_depreciation is not None:
                valuation_data['ExtraordinaryDepreciationMethod'] = extraordinary_depreciation.find('nsSAFT:ExtraordinaryDepreciationMethod', namespaces).text if extraordinary_depreciation.find('nsSAFT:ExtraordinaryDepreciationMethod', namespaces) is not None else ''
                valuation_data['ExtraordinaryDepreciationAmountForPeriod'] = float(extraordinary_depreciation.find('nsSAFT:ExtraordinaryDepreciationAmountForPeriod', namespaces).text) if extraordinary_depreciation.find('nsSAFT:ExtraordinaryDepreciationAmountForPeriod', namespaces) is not None else 0
            
            # Combine asset_data and valuation_data into a single dictionary
            combined_data = {**asset_data, **valuation_data}
            
            # Append the combined data to the assets_data list
            assets.append(combined_data)

    # Convert the list to a DataFrame
    assets_df = pd.DataFrame(assets)

    asset_transactions = []

    for transaction in root.findall('.//nsSAFT:AssetTransaction', namespaces):
        transaction_data = {
            'AssetTransactionID': transaction.find('nsSAFT:AssetTransactionID', namespaces).text if transaction.find('nsSAFT:AssetTransactionID', namespaces) is not None else '',
            'AssetID': transaction.find('nsSAFT:AssetID', namespaces).text if transaction.find('nsSAFT:AssetID', namespaces) is not None else '',
            'AssetTransactionType': transaction.find('nsSAFT:AssetTransactionType', namespaces).text if transaction.find('nsSAFT:AssetTransactionType', namespaces) is not None else '',
            'AssetTransactionDate': transaction.find('nsSAFT:AssetTransactionDate', namespaces).text if transaction.find('nsSAFT:AssetTransactionDate', namespaces) is not None else '',
            'TransactionID': transaction.find('nsSAFT:TransactionID', namespaces).text if transaction.find('nsSAFT:TransactionID', namespaces) is not None else '',
        }
        
        valuations = []
        for valuation in transaction.findall('nsSAFT:AssetTransactionValuations/nsSAFT:AssetTransactionValuation', namespaces):
            valuation_data = {
                'AcquisitionAndProductionCostsOnTransaction': float(valuation.find('nsSAFT:AcquisitionAndProductionCostsOnTransaction', namespaces).text) if valuation.find('nsSAFT:AcquisitionAndProductionCostsOnTransaction', namespaces) is not None else 0,
                'BookValueOnTransaction': float(valuation.find('nsSAFT:BookValueOnTransaction', namespaces).text) if valuation.find('nsSAFT:BookValueOnTransaction', namespaces) is not None else 0,
                'AssetTransactionAmount': float(valuation.find('nsSAFT:AssetTransactionAmount', namespaces).text) if valuation.find('nsSAFT:AssetTransactionAmount', namespaces) is not None else 0
            }
            valuations.append(valuation_data)
        
        combined_data = {**transaction_data, **valuation_data}
            
            # Append the combined data to the assets_data list
        asset_transactions.append(combined_data)

    # Convert the list to a DataFrame
    asset_transactions = pd.DataFrame(asset_transactions)
    return accounts, assets, asset_transactions
#     for sheet_name, columns in summary_data.items():
#         ws[f'A{row}'] = sheet_name
#         formula_parts = [f'SUM(\'{sheet_name}\'!{col}:{col})' for col in columns]
#         formula = f'= {" + ".join(formula_parts)}'
#         ws[f'B{row}'] = formula
#         row += 1

#     # Apply formatting
#     for col_idx in range(1, 3):
#         for row_idx in range(5, row):
#             cell = ws.cell(row=row_idx, column=col_idx)
#             cell.number_format = '#,##0.00'
#             cell.font = Font(size=11)
#             cell.alignment = Alignment(horizontal='right', vertical='center')

#     # Set column width
#     for col_idx in range(1, 3):
#         max_length = max(len(str(cell.value)) for cell in ws[get_column_letter(col_idx)])
#         adjusted_width = (max_length + 2) * 1.2
#         ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

#     # Apply background color
#     for row in ws.iter_rows(min_row=5, max_row=row-1, min_col=1, max_col=2):
#         for cell in row:
#             cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

#     # Merge cells for title
#     ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=2)

# Restul codului rămâne neschimbat

# Extragerea datelor din XML și salvarea într-un fișier Excel
xml_file = 'SAF-T FLINT 2.2023.xml'  # Înlocuiește cu calea către fișierul tău XML
xml_file_Assets = 'SAF-T Assets.xml'
output_file = 'SAF-T FLINT 2.xlsx'
output_file_Assets="Assets.xlsx"
client_name = "EXPEDITORS"  # Înlocuiește cu numele clientului

# accounts, customers, suppliers, invoices, invoices_sales, payments, je = parse_xml_to_dict(xml_file)
# save_to_excel(accounts, customers, suppliers, invoices, invoices_sales, payments, je, client_name, output_file)
# accounts,assets, asset_transactions = parse_xmlAssets_to_dict(xml_file_Assets)
# save_to_excel_Assets(accounts,assets, asset_transactions, output_file_Assets)
# print(f"Informațiile au fost extrase și salvate în {output_file}")